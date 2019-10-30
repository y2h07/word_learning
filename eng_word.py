# -*- coding: utf-8 -*-
"""
Created on Mon Oct 21 17:16:17 2019

@author: lcy
"""

import openpyxl
import random
import sys

def load_xl(wb,row,col):
    word_list = []
    ws = wb.active
    cell_1 = ws.cell(row,col)
    cell_2 = ws.cell(row,col+1)
    word_list.append(cell_1.value)
    c_list = cell_2.value.split(',')
    if len(c_list) == 1:
        word_list.append(c_list[0])
    else:
        word_list.append(c_list[0])
        word_list.append(c_list[1])
    
    return word_list

def eq(w_list,Cword):
    if len(w_list) == 2 :
        if Cword == w_list[1]:
            print("You are right!")
        elif Cword == "quit":
            sys.exit(0)
        else:
            print("You are wrong!Try again!")
            Cword = input("输入中文意思：")
            eq(w_list,Cword)
    elif len(w_list) == 3:
            if Cword == w_list[1] or Cword == w_list[2]:
                print("You are right!")
            elif Cword == "quit":
                sys.exit(0)
            else:
                print("You are wrong!Try again!")
                Cword = input("输入中文意思：")
                eq(w_list,Cword)


if __name__ =="__main__":
    wb = openpyxl.load_workbook("word_list.xlsx")
    
    while True:
        j = random.randint(1,2)   #步长设为2
        if j==1:
            i = random.randint(1,407)
            k = j
        else:
            i = random.randint(1,175)
            k = j + 1
        
        w_list = load_xl(wb,i,k)
        print("英文单词：" + w_list[0])
        Cword = input("输入中文意思：")
        eq(w_list,Cword)
       
        
            
            

    